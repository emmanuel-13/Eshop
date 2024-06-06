import openpyxl

class Excel:
    def icon_function():
        wb = openpyxl.load_workbook('excel/icons.xlsx')
        sheet = wb.active
        return sheet

    choices_icon = []
    for rows in icon_function().iter_rows(values_only=True):
        color_name, color_code = rows
        choices_icon.append((color_name, color_code))
        
        
    def stack_function():
        wb = openpyxl.load_workbook('excel/stack.xlsx')
        sheet = wb.active
        return sheet
    
    stack_icon = []
    for rows in stack_function().iter_rows(values_only=True):
        color_name, color_code = rows
        stack_icon.append((color_name, color_code))
    

    def color_function():
        wb = openpyxl.load_workbook('excel/colors.xlsx')
        sheet = wb.active
        return sheet

    choices_color = []
    for rows in color_function().iter_rows(values_only=True):
        color_name, color_code = rows
        choices_color.append((color_name, color_code))
    